import pandas as pd
import os
import time
import random
import google.generativeai as genai
from google.api_core import exceptions
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ==========================================
# 🔑 設定區
# ==========================================
GEMINI_API_KEY = "AIzaSyDslHJzKaVPVvHJU2f41ix0ebCNQyuZrjQ"

# --- 路徑設定 ---
current_dir = os.path.dirname(os.path.abspath(__file__))
BASE_PATH = os.path.dirname(current_dir)
INPUT_FILE = os.path.join(BASE_PATH, "02_ProcessedData", "BI_Tables", "Dim_Customer.parquet")
CONFIG_FILE = os.path.join(BASE_PATH, "99_Config", "Customer_Parent_Mapping.xlsx")

genai.configure(api_key=GEMINI_API_KEY)
model = genai.GenerativeModel('gemini-1.5-flash')

# ==========================================
# 🛠️ 輔助函式
# ==========================================

def get_learning_examples(df_exist):
    """ [機器學習] 從既有 Excel 抽取範例 (優先找 Manual) """
    examples = []
    if df_exist.empty: return ""
    if 'Category' not in df_exist.columns: return ""
    if 'Source' not in df_exist.columns: df_exist['Source'] = ''

    for cat in ['OEM', 'SI', 'EMS', 'Education', 'Government']:
        manual_rows = df_exist[
            (df_exist['Category'] == cat) & 
            (df_exist['Source'].str.contains('Manual', case=False, na=False))
        ]
        if not manual_rows.empty:
            candidate_rows = manual_rows
        else:
            candidate_rows = df_exist[df_exist['Category'] == cat]

        if not candidate_rows.empty:
            samples = candidate_rows['Original_CustName'].dropna().unique()
            if len(samples) > 0:
                picks = random.sample(list(samples), min(3, len(samples)))
                for p in picks:
                    examples.append(f"- {p}: {cat}")
    return "\n".join(examples)

def apply_hard_rules(name):
    """ [第一層防禦] 硬規則 """
    name_upper = str(name).upper()
    si_keywords = ['LEIDOS', 'GDIT', 'CACI', 'SAIC', 'BOOZ ALLEN', 'AIC ', 'RAICAM', 'LOCKHEED', 'RAYTHEON', 'NORTHROP', 'L3HARRIS'] 
    if any(x in name_upper for x in si_keywords): return 'SI'
    oem_keywords = ['SPACEX', 'TESLA', 'BOEING', 'HONEYWELL', 'GE ', 'GENERAL ELECTRIC', 'SIEMENS', 'SCHNEIDER', 'ABB', 'EATON']
    if any(x in name_upper for x in oem_keywords): return 'OEM'
    if any(x in name_upper for x in ['UNIVERSITY', 'COLLEGE', 'SCHOOL', 'INSTITUTE']): return 'Education'
    if any(x in name_upper for x in ['GOVERNMENT', 'CITY OF', 'STATE OF', 'DEPT OF']): return 'Government'
    if any(x in name_upper for x in ['HOSPITAL', 'MEDICAL', 'CLINIC']): return 'Healthcare'
    return None

def ask_gemini_batch(names_list, learning_text=""):
    """ 
    [雲端大腦] 呼叫 Gemini (快速失敗版) 
    不再重試，遇到 429 直接拋出異常讓主程式接手
    """
    if not names_list: return {}
    
    prompt = f"""
    你是 B2B 產業分析專家。請參考以下的「已知範例」，判斷新公司的 Category (產業分類) 與 Group (集團名稱)。
    【已知範例】：
    {learning_text}
    【分類標準】：
    - OEM: 製造商
    - SI: 系統整合商
    - EMS: 電子代工
    - Education: 學校
    - Government: 政府
    - Distributor: 經銷商
    - Uncategorized: 無法判斷
    請回傳純文字，每行一筆，使用 | 分隔:
    Original_Name|Category|Group_Name
    【待處理名單】：
    {chr(10).join(names_list)}
    """
    
    try:
        response = model.generate_content(prompt)
        text_resp = response.text
        
        result_map = {}
        for line in text_resp.split('\n'):
            if '|' in line and 'Original_Name' not in line:
                parts = line.split('|')
                if len(parts) >= 3:
                    orig = parts[0].strip()
                    cat = parts[1].strip()
                    group = parts[2].strip()
                    result_map[orig] = {'Category': cat, 'Group': group}
        return result_map
        
    except exceptions.ResourceExhausted:
        # 直接往上拋出，不要在這裡睡覺等待
        raise 
    except Exception as e:
        print(f"     ⚠️ API 未知錯誤 (跳過此批): {e}")
        return {}

def apply_excel_formatting_advanced(file_path, df):
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        last_row = len(df) + 1
        last_col = len(df.columns)
        ref = f"A1:{get_column_letter(last_col)}{last_row}"
        tab = Table(displayName="CustomerMapping", ref=ref)
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        try: ws.add_table(tab)
        except ValueError: pass

        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 50)

        # 下拉選單
        cat_options = '"OEM,SI,EMS,Education,Government,Distributor,Healthcare,Uncategorized,Other"'
        dv_cat = DataValidation(type="list", formula1=cat_options, allow_blank=True)
        dv_cat.add(f'C2:C50000') 
        ws.add_data_validation(dv_cat)

        source_options = '"Manual,Gemini-AI,Hard-Rule,Check-Manually"'
        dv_source = DataValidation(type="list", formula1=source_options, allow_blank=True)
        dv_source.add(f'D2:D50000')
        ws.add_data_validation(dv_source)

        wb.save(file_path)
    except Exception:
        pass

# ==========================================
# 🚀 主程式 (V7.6 快速失敗版)
# ==========================================

def run_detective():
    print("🕵️‍♂️ [集團偵查兵 V7.6 - 快速離線版] 啟動中...")
    
    if not os.path.exists(INPUT_FILE):
        print(f"❌ 找不到輸入檔: {INPUT_FILE}")
        return
    df_cust = pd.read_parquet(INPUT_FILE)
    all_customers = df_cust[['CustName']].drop_duplicates()
    
    # 讀取 Excel
    if os.path.exists(CONFIG_FILE):
        print("   - 讀取既有帳本...")
        df_exist = pd.read_excel(CONFIG_FILE)
        # 更新欄位
        rename_dict = {'Tag': 'Category', 'Note': 'Source'}
        df_exist = df_exist.rename(columns=rename_dict)
        for col in ['Original_CustName', 'Parent_Group', 'Category', 'Source']:
            if col not in df_exist.columns: df_exist[col] = ''
        learning_examples = get_learning_examples(df_exist)
    else:
        print("   - 建立新帳本...")
        df_exist = pd.DataFrame(columns=['Original_CustName', 'Parent_Group', 'Category', 'Source'])
        learning_examples = ""

    processed_set = set(df_exist['Original_CustName'])
    target_customers = all_customers[~all_customers['CustName'].isin(processed_set)]
    
    if len(target_customers) == 0:
        print("✅ 所有客戶都已在帳本中，無需更新。")
        return

    print(f"   - 發現 {len(target_customers):,} 筆新客戶...")
    
    new_results = []
    batch_for_ai = [] 
    
    # Phase 1: 硬規則 (本機運算，超快)
    print("   - Phase 1: 硬規則過濾...")
    for idx, row in target_customers.iterrows():
        orig_name = row['CustName']
        hard_cat = apply_hard_rules(orig_name)
        if hard_cat:
            group_name = orig_name
            if 'SPACEX' in orig_name.upper(): group_name = 'SPACEX GROUP'
            if 'GDIT' in orig_name.upper(): group_name = 'GDIT GROUP'
            if 'LEIDOS' in orig_name.upper(): group_name = 'LEIDOS GROUP'
            
            new_results.append({
                'Original_CustName': orig_name,
                'Parent_Group': group_name,
                'Category': hard_cat,
                'Source': 'Hard-Rule'
            })
        else:
            batch_for_ai.append(orig_name)

    # Phase 2: Gemini API (帶有快速熔斷機制)
    if batch_for_ai:
        print(f"   - Phase 2: Gemini API 判讀 (批次處理)...")
        batch_size = 20 
        total_batches = (len(batch_for_ai) + batch_size - 1) // batch_size
        
        api_quota_exhausted = False # 標記 API 是否掛了

        for i in range(0, len(batch_for_ai), batch_size):
            batch = batch_for_ai[i:i+batch_size]
            current_batch_idx = i//batch_size + 1
            
            ai_results = {}
            
            # 如果 API 還活著，就試著呼叫
            if not api_quota_exhausted:
                print(f"     Batch {current_batch_idx}/{total_batches} (AI)...")
                try:
                    ai_results = ask_gemini_batch(batch, learning_examples)
                    time.sleep(2) # 正常休息
                except exceptions.ResourceExhausted:
                    print("     ⚠️ API 額度用盡 (429)！切換至離線模式，後續將跳過 AI。")
                    api_quota_exhausted = True # 觸發熔斷
                except Exception as e:
                    print(f"     ⚠️ API 錯誤: {e}，跳過此批。")
            else:
                # 離線模式，不印 Log 或只印進度
                if current_batch_idx % 10 == 0: # 減少洗版
                    print(f"     Batch {current_batch_idx}/{total_batches} (Offline)...")

            # 填寫結果
            for name in batch:
                final_cat = 'Uncategorized'
                final_group = name
                source = 'Check-Manually'
                
                if name in ai_results:
                    res = ai_results[name]
                    if res['Category'] and res['Category'].lower() != 'other' and res['Category'] != '':
                        final_cat = res['Category']
                        final_group = res['Group']
                        source = 'Gemini-AI'
                
                new_results.append({
                    'Original_CustName': name,
                    'Parent_Group': final_group,
                    'Category': final_cat,
                    'Source': source
                })

    # 存檔
    if new_results:
        new_df = pd.DataFrame(new_results)
        final_df = pd.concat([df_exist, new_df], ignore_index=True)
        final_df['Category'] = final_df['Category'].fillna('Uncategorized')
        final_df.loc[final_df['Category'] == '', 'Category'] = 'Uncategorized'
        final_df['Parent_Group'] = final_df['Parent_Group'].fillna(final_df['Original_CustName'])
        
        final_df = final_df.sort_values(by=['Source', 'Original_CustName'], ascending=[True, True])
        
        final_df.to_excel(CONFIG_FILE, index=False)
        apply_excel_formatting_advanced(CONFIG_FILE, final_df)
        print(f"✨ 已更新帳本: {CONFIG_FILE} (新增 {len(new_results)} 筆)")
    else:
        print("✨ 暫無新資料需更新。")

if __name__ == "__main__":
    run_detective()